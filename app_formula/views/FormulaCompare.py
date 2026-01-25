from django.views.generic import TemplateView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from app_repository.models import MaterialLibrary, TestConfig
from app_formula.models import LabFormula
from app_raw_material.models import RawMaterial
from collections import defaultdict
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import json

# ==========================================
# 1. 对比购物车 API
# ==========================================
class FormulaCompareCartView(LoginRequiredMixin, View):
    """
    处理对比列表的增删改查 (基于 Session)
    Session Key: 'cart_formulas_v2' -> [id1, id2, ...] (V2版本，彻底隔离)
    Session Key: 'cart_materials_v2' -> [id1, id2, ...]
    Session Key: 'cart_raw_materials_v2' -> [id1, id2, ...] (新增：原材料)
    """
    
    def get(self, request):
        """获取当前对比列表"""
        formula_ids = request.session.get('cart_formulas_v2', [])
        material_ids = request.session.get('cart_materials_v2', [])
        raw_material_ids = request.session.get('cart_raw_materials_v2', [])
        
        formulas = LabFormula.objects.filter(pk__in=formula_ids).values('id', 'code', 'name')
        materials = MaterialLibrary.objects.filter(pk__in=material_ids).values('id', 'grade_name', 'manufacturer')
        raw_materials = RawMaterial.objects.filter(pk__in=raw_material_ids).values('id', 'name', 'model_name', 'supplier__name')
        
        # 统一格式返回
        items = []
        for f in formulas:
            items.append({'id': f['id'], 'name': f['code'], 'desc': f['name'], 'type': 'formula'})
        for m in materials:
            items.append({'id': m['id'], 'name': m['grade_name'], 'desc': m['manufacturer'], 'type': 'material'})
        for rm in raw_materials:
            desc = rm['supplier__name'] if rm['supplier__name'] else ''
            name = f"{rm['name']} {rm['model_name']}" if rm['model_name'] else rm['name']
            items.append({'id': rm['id'], 'name': name, 'desc': desc, 'type': 'raw_material'})
            
        return JsonResponse({'count': len(items), 'items': items})

    def post(self, request):
        """加入/移除对比"""
        action = request.POST.get('action') # 'add', 'remove', 'clear', 'toggle', 'add_multiple'
        item_type = request.POST.get('type') # 'formula' or 'material' or 'raw_material'
        
        # 【严格检查】必须指定 type，且必须合法
        if action != 'clear' and item_type not in ['formula', 'material', 'raw_material']:
            return JsonResponse({'status': 'error', 'message': 'Invalid type parameter'}, status=400)
        
        # 使用 list() 创建副本，防止引用问题
        formula_ids = list(request.session.get('cart_formulas_v2', []))
        material_ids = list(request.session.get('cart_materials_v2', []))
        raw_material_ids = list(request.session.get('cart_raw_materials_v2', []))
        
        # 确定目标列表
        if item_type == 'material':
            target_list = material_ids
        elif item_type == 'raw_material':
            target_list = raw_material_ids
        else:
            target_list = formula_ids
        
        if action == 'clear':
            # 清空所有
            formula_ids = []
            material_ids = []
            raw_material_ids = []
            target_list = [] # 重置引用
            
        elif action == 'add_multiple':
            # 批量添加
            ids_str = request.POST.get('ids') # JSON string or comma separated
            if ids_str:
                try:
                    # 尝试解析 JSON 列表
                    new_ids = json.loads(ids_str)
                    if not isinstance(new_ids, list):
                        new_ids = [int(ids_str)]
                except:
                    # 尝试解析逗号分隔
                    new_ids = [int(x) for x in ids_str.split(',') if x.isdigit()]
                
                for fid in new_ids:
                    try:
                        fid_int = int(fid)
                        if fid_int not in target_list:
                            target_list.append(fid_int)
                    except (ValueError, TypeError):
                        continue
                        
        else:
            # 单个操作
            item_id = request.POST.get('id')
            if item_id:
                try:
                    fid = int(item_id)
                    if action == 'add':
                        if fid not in target_list:
                            target_list.append(fid)
                    elif action == 'remove':
                        if fid in target_list:
                            target_list.remove(fid)
                    elif action == 'toggle':
                        if fid in target_list:
                            target_list.remove(fid)
                        else:
                            target_list.append(fid)
                except ValueError:
                    pass
        
        # 更新 Session
        if action == 'clear':
            request.session['cart_formulas_v2'] = []
            request.session['cart_materials_v2'] = []
            request.session['cart_raw_materials_v2'] = []
        else:
            if item_type == 'material':
                request.session['cart_materials_v2'] = target_list
            elif item_type == 'raw_material':
                request.session['cart_raw_materials_v2'] = target_list
            elif item_type == 'formula':
                request.session['cart_formulas_v2'] = target_list
            
        total_count = len(request.session.get('cart_formulas_v2', [])) + \
                      len(request.session.get('cart_materials_v2', [])) + \
                      len(request.session.get('cart_raw_materials_v2', []))
        
        return JsonResponse({
            'status': 'success', 
            'count': total_count, 
            'ids': target_list,
            'type': item_type
        })


# ==========================================
# 2. 对比页面视图
# ==========================================
class FormulaCompareView(LoginRequiredMixin, TemplateView):
    template_name = 'apps/app_formula/compare.html'

    def post(self, request, *args, **kwargs):
        # 检查是否是导出请求
        if 'export_excel' in request.POST:
            return self.export_excel(request)
        return self.get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # 1. 获取参数
        # 支持 POST (表单提交) 和 GET (URL参数)
        if self.request.method == 'POST':
            material_id = self.request.POST.get('material_id') # 单个基准材料 (旧逻辑兼容)
            # 过滤空值，防止 [''] 导致 int 转换错误
            formula_ids = [x for x in self.request.POST.getlist('formula_ids') if x]
            material_ids = [x for x in self.request.POST.getlist('material_ids') if x]
            raw_material_ids = [x for x in self.request.POST.getlist('raw_material_ids') if x]

            # 如果完全没传参数，回退到 Session
            if not formula_ids and not material_ids and not raw_material_ids and not material_id:
                formula_ids = self.request.session.get('cart_formulas_v2', [])
                material_ids = self.request.session.get('cart_materials_v2', [])
                raw_material_ids = self.request.session.get('cart_raw_materials_v2', [])

        else:
            material_id = self.request.GET.get('material_id')
            formula_ids = self.request.GET.getlist('ids') # 优先从 URL 获取 ids (旧逻辑兼容)
            material_ids = self.request.GET.getlist('m_ids')
            raw_material_ids = self.request.GET.getlist('rm_ids')
            
            # 如果 URL 没参数，从 Session 获取
            if not formula_ids and not material_ids and not raw_material_ids and not material_id:
                formula_ids = self.request.session.get('cart_formulas_v2', [])
                material_ids = self.request.session.get('cart_materials_v2', [])
                raw_material_ids = self.request.session.get('cart_raw_materials_v2', [])

        # 2. 获取对象
        formulas = list(LabFormula.objects.filter(pk__in=formula_ids).order_by('created_at'))
        materials = list(MaterialLibrary.objects.filter(pk__in=material_ids).order_by('created_at'))
        raw_materials = list(RawMaterial.objects.filter(pk__in=raw_material_ids).order_by('created_at'))
        
        # 兼容旧的单基准材料逻辑
        base_material = None
        if material_id:
            base_material = get_object_or_404(MaterialLibrary, pk=material_id)
            # 如果基准材料不在列表里，加进去作为第一列
            if base_material not in materials:
                materials.insert(0, base_material)
        
        if not formulas and not materials and not raw_materials:
             messages.warning(self.request, "请先选择要对比的项目")
             return context

        # 3. 定义列头 (混合排序或分组)
        # 策略：先放材料，再放原材料，再放配方
        columns = []
        for m in materials:
            columns.append({'type': 'material', 'obj': m})
        for rm in raw_materials:
            columns.append({'type': 'raw_material', 'obj': rm})
        for f in formulas:
            columns.append({'type': 'formula', 'obj': f})

        # ==========================================
        # 4. 构建 BOM 对比矩阵 (仅配方有 BOM)
        # ==========================================
        bom_matrix = []
        all_raw_materials = set()
        for f in formulas:
            for line in f.bom_lines.all():
                all_raw_materials.add(line.raw_material)
        
        sorted_raw_materials = sorted(list(all_raw_materials), key=lambda x: (x.category.order, x.name))

        for rm in sorted_raw_materials:
            row = {
                'item': rm,
                'values': []
            }
            
            for col in columns:
                if col['type'] == 'formula':
                    f = col['obj']
                    line = f.bom_lines.filter(raw_material=rm).first()
                    if line:
                        row['values'].append({'val': line.percentage, 'is_highlight': True})
                    else:
                        row['values'].append({'val': '-', 'is_highlight': False})
                else:
                    # 材料和原材料没有 BOM
                    row['values'].append({'val': '-', 'is_highlight': False})
            bom_matrix.append(row)

        # ==========================================
        # 5. 构建 性能对比矩阵 (混合)
        # ==========================================
        test_matrix = []
        all_test_configs = set()
        
        # 收集所有涉及的测试配置
        for m in materials:
            for p in m.properties.all():
                all_test_configs.add(p.test_config)
        
        for rm in raw_materials:
            for p in rm.properties.all():
                all_test_configs.add(p.test_config)
            
        for f in formulas:
            for res in f.test_results.all():
                all_test_configs.add(res.test_config)
                
        sorted_configs = sorted(list(all_test_configs), key=lambda x: (x.category.order, x.order))
        
        # 预加载数据
        mat_props = {} # {mat_id: {config_id: val}}
        for m in materials:
            # 兼容非数值类型
            mat_props[m.id] = {
                p.test_config_id: p.value_text if p.test_config.data_type != 'NUMBER' else p.value 
                for p in m.properties.all()
            }
            
        raw_mat_props = {} # {raw_mat_id: {config_id: val}}
        for rm in raw_materials:
            raw_mat_props[rm.id] = {
                p.test_config_id: p.value_text if p.test_config.data_type != 'NUMBER' else p.value 
                for p in rm.properties.all()
            }
            
        formula_props = {} # {formula_id: {config_id: val}}
        for f in formulas:
            formula_props[f.id] = {
                r.test_config_id: r.value_text if r.test_config.data_type != 'NUMBER' else r.value 
                for r in f.test_results.all()
            }

        # 确定基准值 (取第一列的值作为基准)
        first_col = columns[0] if columns else None
        
        for tc in sorted_configs:
            row = {
                'item': tc,
                'values': []
            }
            
            # 获取基准值
            base_val = None
            if first_col:
                if first_col['type'] == 'material':
                    base_val = mat_props.get(first_col['obj'].id, {}).get(tc.id)
                elif first_col['type'] == 'raw_material':
                    base_val = raw_mat_props.get(first_col['obj'].id, {}).get(tc.id)
                else:
                    base_val = formula_props.get(first_col['obj'].id, {}).get(tc.id)
            
            # 填充数据
            for i, col in enumerate(columns):
                val = None
                if col['type'] == 'material':
                    val = mat_props.get(col['obj'].id, {}).get(tc.id)
                elif col['type'] == 'raw_material':
                    val = raw_mat_props.get(col['obj'].id, {}).get(tc.id)
                else:
                    val = formula_props.get(col['obj'].id, {}).get(tc.id)
                
                # 对比逻辑 (从第二列开始对比)
                compare_class = ""
                # 只有数值类型才进行大小比较
                if i > 0 and val is not None and base_val is not None and tc.data_type == 'NUMBER':
                    try:
                        if val > base_val:
                            compare_class = "text-green"
                        elif val < base_val:
                            compare_class = "text-red"
                    except: pass
                
                row['values'].append({
                    'val': val if val is not None else '-',
                    'compare_class': compare_class,
                    'is_base': (i == 0)
                })
                
            test_matrix.append(row)

        context['columns'] = columns
        context['bom_matrix'] = bom_matrix
        context['test_matrix'] = test_matrix
        context['page_title'] = "综合对比分析"
        # 传递 material 对象以便模板兼容旧逻辑 (如果有且仅有一个材料且在第一位)
        if materials and len(materials) == 1 and columns[0]['type'] == 'material':
            context['material'] = materials[0]
        
        return context

    def export_excel(self, request):
        """导出 Excel 报表"""
        context = self.get_context_data()
        if not context or 'columns' not in context:
            return redirect('formula_list')
            
        columns = context['columns']
        bom_matrix = context['bom_matrix']
        test_matrix = context['test_matrix']
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "对比报表"
        
        # 样式
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        red_font = Font(color="FF0000", bold=True)
        green_font = Font(color="008000", bold=True)
        orange_font = Font(color="FFA500", bold=True)
        
        # 1. 表头
        # 修改表头结构：拆分第一列
        headers = ["分类 / 项目", "详情 / 标准", "单位"]
        for col in columns:
            if col['type'] == 'material':
                headers.append(f"材料\n{col['obj'].grade_name}")
            elif col['type'] == 'raw_material':
                name = f"{col['obj'].name} {col['obj'].model_name}" if col['obj'].model_name else col['obj'].name
                headers.append(f"原材料\n{name}")
            else:
                headers.append(f"配方\n{col['obj'].code}\n{col['obj'].name}")
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
            
        # 2. 描述
        desc_row = ["描述/备注", "-", "-"]
        for col in columns:
            if col['type'] == 'material':
                 desc_row.append(col['obj'].description or "-")
            elif col['type'] == 'raw_material':
                 desc_row.append(col['obj'].usage_method or "-")
            else:
                 desc_row.append(col['obj'].description or "-")
        ws.append(desc_row)
        
        # 3. 成本
        pred_cost_row = ["预测成本", "-", "元/kg"]
        for col in columns:
            if col['type'] == 'formula':
                 pred_cost_row.append(col['obj'].cost_predicted)
            elif col['type'] == 'raw_material':
                 pred_cost_row.append(col['obj'].cost_price or "-")
            else:
                 pred_cost_row.append("-")
        ws.append(pred_cost_row)
        
        act_cost_row = ["实际成本", "-", "元/kg"]
        for col in columns:
            if col['type'] == 'formula':
                 act_cost_row.append(col['obj'].cost_actual or "-")
            else:
                 act_cost_row.append("-")
        ws.append(act_cost_row)
        
        # 4. BOM
        ws.append(["BOM 结构对比"])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len(headers))
        ws.cell(row=ws.max_row, column=1).fill = PatternFill(start_color="FFF5E6", end_color="FFF5E6", fill_type="solid")
        ws.cell(row=ws.max_row, column=1).font = orange_font
        
        for row in bom_matrix:
            # 修改 BOM 行结构
            # 第一列：原材料类型
            # 第二列：原材料名称 + 型号
            # 第三列：单位
            item_name = f"{row['item'].name} {row['item'].model_name or ''}"
            data_row = [row['item'].category.name, item_name, "%"]
            for cell in row['values']:
                data_row.append(cell['val'])
            ws.append(data_row)
            
        # 5. 性能
        ws.append(["性能指标对比"])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len(headers))
        ws.cell(row=ws.max_row, column=1).fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
        ws.cell(row=ws.max_row, column=1).font = Font(color="800080", bold=True)
        
        for row in test_matrix:
            # 修改性能行结构
            # 第一列：指标名称
            # 第二列：标准 + 条件
            # 第三列：单位
            standard_info = row['item'].standard
            if row['item'].condition:
                standard_info += f" ({row['item'].condition})"
                
            data_row = [row['item'].name, standard_info, row['item'].unit]
            ws.append(data_row + [c['val'] for c in row['values']])
            current_row_idx = ws.max_row
            
            # 颜色标记 (从第4列开始，因为前3列是固定列)
            for i, cell in enumerate(row['values']):
                if cell.get('compare_class') == 'text-green':
                    ws.cell(row=current_row_idx, column=i+4).font = green_font
                elif cell.get('compare_class') == 'text-red':
                    ws.cell(row=current_row_idx, column=i+4).font = red_font
                    
        # 样式调整
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = center_align
                cell.border = border
                if cell.column == 1 or cell.column == 2: # 前两列左对齐
                    cell.alignment = left_align
                    
        ws.column_dimensions['A'].width = 20 # 分类/项目
        ws.column_dimensions['B'].width = 30 # 详情/标准
        ws.column_dimensions['C'].width = 10 # 单位
        for i in range(4, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 20
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"综合对比报表.xlsx"
        # 处理中文文件名
        from django.utils.encoding import escape_uri_path
        response['Content-Disposition'] = f'attachment; filename="{escape_uri_path(filename)}"'
        
        wb.save(response)
        return response
