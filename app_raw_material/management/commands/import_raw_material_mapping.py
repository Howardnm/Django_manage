"""
Django 管理命令: SAP 原材料清洗与类型关联（Excel 往返）。

背景:
    SAP 同步 (sync_raw_materials) 把物料描述 MAKTX 整体塞进 RawMaterial.name,
    category 固定为 "未分类"。本命令提供一次性的批量清洗底稿:

    1. --generate: 导出全部 warehouse_code 非空的原材料到 init/raw_material_mapping.xlsx,
       文员在 Excel 中拆分名称/型号, 并通过下拉列表选定类型。
    2. (无参数): 读取底稿, 按 sap_code 比对 RawMaterial.warehouse_code 回写
       name / model_name / category。

用法:
    python manage.py import_raw_material_mapping --generate   # 导出/更新底稿
    python manage.py import_raw_material_mapping              # 导入回写
    python manage.py import_raw_material_mapping --dry-run    # 导入前预览
"""

import os

from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from app_raw_material.models import RawMaterial, RawMaterialType

WORKBOOK_PATH = os.path.join(settings.BASE_DIR, 'init', 'raw_material_mapping.xlsx')

MAIN_SHEET = "清洗底稿"
REF_SHEET = "类型参考"

# 列顺序 (1-indexed): A sap_code, B original_name, C original_model, D name, E model_name, F category_name
HEADERS = ["sap_code", "original_name", "original_model", "name", "model_name", "category_name"]
READONLY_COLS = {1, 2, 3}  # A/B/C 只读参考列
REF_COL = 6                # F 列: category_name 下拉

# 数据验证覆盖行数 (预留增量空行)
DATA_VALIDATION_ROWS = 2000

# 下拉显示值 = "名称 || 描述" 的分隔符 (避开描述内容, 方便导入时解析)
SEP = " || "


class Command(BaseCommand):
    help = "SAP 原材料清洗与类型关联：--generate 导出 xlsx 底稿，无参数导入回写"

    def add_arguments(self, parser):
        parser.add_argument(
            "--generate", action="store_true",
            help="导出/更新 init/raw_material_mapping.xlsx 工作底稿",
        )
        parser.add_argument(
            "--dry-run", action="store_true",
            help="导入前预览，不写入数据库",
        )

    def handle(self, *args, **options):
        if options["generate"]:
            self._generate()
        else:
            self._import(dry_run=options["dry_run"])

    # ------------------------------------------------------------------
    # 公共工具
    # ------------------------------------------------------------------

    @staticmethod
    def _clean_str(value):
        """openpyxl 读出的值规范化: None->'', float 整数->int 字符串, 其余 strip"""
        if value is None:
            return ""
        if isinstance(value, float):
            return str(int(value)) if value.is_integer() else str(value)
        return str(value).strip()

    @classmethod
    def _format_category(cls, name, desc):
        """生成下拉显示值: 有描述时 '名称 || 描述', 否则 '名称'。
        描述中的 '||' 替换为全角 '｜', 保证分隔符唯一。"""
        if desc:
            desc_clean = desc.replace(SEP, "｜")
            return f"{name}{SEP}{desc_clean}"
        return name

    @classmethod
    def _parse_category(cls, value):
        """从下拉显示值解析出类型名称: 取 ' || ' 前段; 无分隔符则整体视为名称。"""
        if SEP in value:
            return value.split(SEP, 1)[0].strip()
        return value.strip()

    # ------------------------------------------------------------------
    # 导出 (--generate)
    # ------------------------------------------------------------------

    def _generate(self):
        self.stdout.write(
            self.style.MIGRATE_HEADING("\n=== 生成/更新 SAP 原材料清洗底稿 ... ===")
        )

        # 1. 读取旧底稿中已填写的值 (增量幂等: 保留文员编辑, 不覆盖)
        existing = self._read_existing()

        # 2. 全量取 warehouse_code 非空的原材料
        qs = (
            RawMaterial.objects
            .exclude(warehouse_code__isnull=True)
            .exclude(warehouse_code="")
            .order_by("warehouse_code")
        )

        # 3. 类型下拉数据源 (名称 + 描述)
        type_options = list(
            RawMaterialType.objects.order_by("order", "name")
            .values_list("name", "description")
        )

        wb = Workbook()
        ws = wb.active
        ws.title = MAIN_SHEET

        # 表头
        grey = PatternFill("solid", fgColor="D9D9D9")
        light_grey = PatternFill("solid", fgColor="F2F2F2")
        for col, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = grey
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

        # 数据行
        row = 2
        for rm in qs.iterator():
            code = str(rm.warehouse_code).strip()
            prev = existing.get(code, {})

            # A sap_code (文本格式防前导零丢失)
            a_cell = ws.cell(row=row, column=1, value=code)
            a_cell.number_format = "@"

            # B/C 原始名称/型号参考 (只读)
            ws.cell(row=row, column=2, value=rm.name)
            ws.cell(row=row, column=3, value=rm.model_name or "")

            # D/E/F: 仅当该行已清洗 (选定了类型) 时才保留已编辑值;
            # 未清洗的行 name/model_name 留空, 便于文员识别尚未处理的行
            cleaned = bool(prev.get("category_name"))
            ws.cell(row=row, column=4, value=prev.get("name", "") if cleaned else "")
            ws.cell(row=row, column=5, value=prev.get("model_name", "") if cleaned else "")
            ws.cell(row=row, column=6, value=prev.get("category_name", ""))

            # 只读列浅灰标识
            for col in READONLY_COLS:
                ws.cell(row=row, column=col).fill = light_grey

            row += 1

        # 4. F 列数据验证下拉 (严格模式, 数据源来自隐藏参考页)
        if type_options:
            dv = DataValidation(
                type="list",
                formula1=f"'{REF_SHEET}'!$A:$A",
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="无效的类型",
                error="请从下拉列表中选择有效的原材料类型。",
            )
            ws.add_data_validation(dv)
            dv.add(f"F2:F{DATA_VALIDATION_ROWS}")

        # 5. 隐藏参考页: 类型下拉值 (名称 || 描述)
        ref = wb.create_sheet(REF_SHEET)
        for i, (tname, tdesc) in enumerate(type_options, start=1):
            ref.cell(row=i, column=1, value=self._format_category(tname, tdesc))
        ref.sheet_state = "hidden"

        wb.save(WORKBOOK_PATH)
        self.stdout.write(
            self.style.SUCCESS(
                f"[OK] 已生成/更新 {WORKBOOK_PATH}"
            )
        )
        self.stdout.write(
            f"     共 {row - 1} 条记录, 可供下拉选择 {len(type_options)} 个类型。"
        )

    def _read_existing(self):
        """读取旧底稿中已填写的数据 {sap_code: {name, model_name, category_name}}"""
        if not os.path.exists(WORKBOOK_PATH):
            return {}
        try:
            wb = load_workbook(WORKBOOK_PATH, data_only=True)
        except Exception:
            return {}
        if MAIN_SHEET not in wb.sheetnames:
            return {}
        ws = wb[MAIN_SHEET]

        result = {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or len(r) < len(HEADERS):
                continue
            code = self._clean_str(r[0])
            if not code:
                continue
            result[code] = {
                "name": self._clean_str(r[3]),
                "model_name": self._clean_str(r[4]),
                "category_name": self._clean_str(r[5]),
            }
        return result

    # ------------------------------------------------------------------
    # 导入回写
    # ------------------------------------------------------------------

    def _import(self, dry_run=False):
        if dry_run:
            self.stdout.write(
                self.style.WARNING("\n=== [DRY-RUN] 预览模式，不会写入数据库 ===")
            )
        else:
            self.stdout.write(
                self.style.MIGRATE_HEADING("\n=== 开始导入 SAP 原材料清洗结果 ... ===")
            )

        if not os.path.exists(WORKBOOK_PATH):
            self.stdout.write(self.style.ERROR(f"未找到工作底稿: {WORKBOOK_PATH}"))
            return

        try:
            wb = load_workbook(WORKBOOK_PATH, data_only=True)
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"读取工作簿失败: {e}"))
            return
        if MAIN_SHEET not in wb.sheetnames:
            self.stdout.write(self.style.ERROR(f"工作簿缺少 {MAIN_SHEET} 页"))
            return
        ws = wb[MAIN_SHEET]

        # 预加载映射字典, 避免 N+1
        warehouse_map = {rm.warehouse_code: rm for rm in RawMaterial.objects.all()}
        type_map = {t.name: t for t in RawMaterialType.objects.all()}

        updated = 0
        skipped = {"no_sap": 0, "no_match": 0, "no_name": 0, "no_category": 0, "no_type": 0, "dup": 0}
        seen = set()

        # 先收集所有行
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or len(r) < len(HEADERS):
                continue
            rows.append(
                (
                    self._clean_str(r[0]),  # sap_code
                    self._clean_str(r[3]),  # name
                    self._clean_str(r[4]),  # model_name
                    self._clean_str(r[5]),  # category_name
                )
            )

        with transaction.atomic():
            for code, name, model_name, category_name in rows:
                if not code:
                    skipped["no_sap"] += 1
                    continue
                if code in seen:
                    skipped["dup"] += 1
                    self.stdout.write(self.style.WARNING(f"  [!] 重复 sap_code: {code}"))
                    continue
                seen.add(code)

                rm = warehouse_map.get(code)
                if rm is None:
                    skipped["no_match"] += 1
                    self.stdout.write(self.style.WARNING(f"  [!] 库里无匹配物料: {code}"))
                    continue
                if not name:
                    skipped["no_name"] += 1
                    self.stdout.write(self.style.WARNING(f"  [!] 拆分后名称为空: {code}"))
                    continue
                if not category_name:
                    skipped["no_category"] += 1
                    self.stdout.write(self.style.WARNING(f"  [!] 未选择类型: {code}"))
                    continue

                # 下拉值可能是 "名称 || 描述", 解析出类型名称后精确匹配
                cat_name = self._parse_category(category_name)
                cat = type_map.get(cat_name)
                if cat is None:
                    skipped["no_type"] += 1
                    self.stdout.write(
                        self.style.WARNING(f"  [!] 类型不存在: {category_name} ({code})")
                    )
                    continue

                if dry_run:
                    self.stdout.write(
                        f"  [~] 将更新: {code} | 名称={name} | 型号={model_name} | 类型={category_name}"
                    )
                else:
                    rm.name = name
                    rm.model_name = model_name
                    rm.category = cat
                    rm.save(update_fields=["name", "model_name", "category", "updated_at"])
                updated += 1

        self.stdout.write(self.style.SUCCESS("\n导入完成！"))
        self.stdout.write(
            f"更新 {updated} 条。跳过: "
            f"无编码 {skipped['no_sap']}, 无匹配 {skipped['no_match']}, "
            f"名称为空 {skipped['no_name']}, 类型为空 {skipped['no_category']}, "
            f"类型不存在 {skipped['no_type']}, 重复 {skipped['dup']}。"
        )